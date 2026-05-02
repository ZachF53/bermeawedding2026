import csv
import io
import json
import logging

from django.conf import settings
from django.contrib import messages
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.core.mail import send_mail
from django.db import transaction
from django.db.models import Count, Q, Sum
from django.http import JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.views.decorators.csrf import ensure_csrf_cookie
from django.views.decorators.http import require_GET, require_POST

try:
    import openpyxl
except ImportError:
    openpyxl = None

from .models import (
    StoryMoment,
    Event,
    CouplePhoto,
    InvitedGuest,
    RSVPSubmission,
    RSVPGuest,
    RSVPEventSelection,
)

logger = logging.getLogger(__name__)

DASHBOARD_LOGIN_URL = '/dashboard/login/'


# ----------------------------- Page views -----------------------------

def home(request):
    hero = CouplePhoto.objects.filter(is_hero=True).first()
    photos = CouplePhoto.objects.filter(is_hero=False)[:6]
    return render(request, 'wedding/index.html', {'hero': hero, 'photos': photos})


def our_story(request):
    moments = StoryMoment.objects.all()
    return render(request, 'wedding/our_story.html', {'moments': moments})


def events(request):
    items = Event.objects.all()
    return render(request, 'wedding/events.html', {'events': items})


@ensure_csrf_cookie
def rsvp(request):
    return render(request, 'wedding/rsvp.html', {
        'rsvp_secret': settings.RSVP_SECRET,
    })


def accommodations(request):
    return render(request, 'wedding/accommodations.html')


def faq(request):
    return render(request, 'wedding/faq.html')


# ----------------------------- RSVP API -------------------------------

def _serialize_existing_rsvp(submission):
    return {
        'attending': submission.attending,
        'email': submission.email,
        'phone': submission.phone,
        'notes': submission.notes,
        'guests': [
            {'first_name': g.first_name, 'last_name': g.last_name}
            for g in submission.guests.all()
        ],
        'events': [e.event_name for e in submission.event_selections.all()],
        'submitted_at': submission.submitted_at.isoformat(),
    }


@require_GET
def rsvp_lookup(request):
    email = (request.GET.get('email') or '').strip().lower()
    first_name = (request.GET.get('first_name') or '').strip()
    last_name = (request.GET.get('last_name') or '').strip()

    if not email and not first_name and not last_name:
        return JsonResponse({'found': False})

    if email:
        matches = list(InvitedGuest.objects.filter(email__iexact=email))
    else:
        qs = InvitedGuest.objects.all()
        if first_name:
            qs = qs.filter(first_name__icontains=first_name)
        if last_name:
            qs = qs.filter(last_name__icontains=last_name)
        matches = list(qs)

    if not matches:
        return JsonResponse({'found': False})

    if len(matches) > 1:
        return JsonResponse({
            'found': True,
            'multiple': True,
            'candidates': [
                {
                    'id': g.id,
                    'full_name': f'{g.first_name} {g.last_name}',
                    'email': g.email or '',
                    'max_guests': g.max_guests,
                    'rsvped': 1 if g.rsvped else 0,
                }
                for g in matches
            ],
        })

    g = matches[0]
    last_submission = g.submissions.order_by('-submitted_at').first()
    return JsonResponse({
        'found': True,
        'multiple': False,
        'invite_id': g.id,
        'email': g.email or '',
        'full_name': f'{g.first_name} {g.last_name}',
        'max_guests': g.max_guests,
        'current_guests': g.current_guests,
        'rsvped': 1 if g.rsvped else 0,
        'rsvp': _serialize_existing_rsvp(last_submission) if last_submission else None,
    })


@require_POST
def rsvp_submit(request):
    try:
        payload = json.loads(request.body or b'{}')
    except json.JSONDecodeError:
        return JsonResponse({'ok': False, 'error': 'Invalid JSON'}, status=400)

    if (payload.get('hp') or '').strip():
        return JsonResponse({'ok': False, 'error': 'Spam detected'}, status=400)

    if payload.get('secret') != settings.RSVP_SECRET:
        return JsonResponse({'ok': False, 'error': 'Invalid request'}, status=403)

    guests = payload.get('guests') or []
    if not isinstance(guests, list) or not guests:
        return JsonResponse({'ok': False, 'error': 'No guests provided'}, status=422)

    contact = payload.get('contact') or {}
    contact_email = (contact.get('email') or '').strip()
    if not contact_email:
        return JsonResponse({'ok': False, 'error': 'Contact email required'}, status=422)

    invited = None
    invite_id = payload.get('invite_id')
    if invite_id:
        invited = InvitedGuest.objects.filter(pk=invite_id).first()
    if not invited:
        invited = InvitedGuest.objects.filter(email__iexact=contact_email).first()
    if not invited:
        return JsonResponse({'ok': False, 'error': 'Invitation not found'}, status=404)

    contact_phone = (contact.get('phone') or '').strip()
    notes = (payload.get('notes') or '').strip()
    top_events = payload.get('topEventSelections') or []

    try:
        with transaction.atomic():
            submission = RSVPSubmission.objects.create(
                invited_guest=invited,
                attending=True,
                email=contact_email,
                phone=contact_phone,
                notes=notes,
            )

            stored_guests = 0
            for g in guests:
                if not isinstance(g, dict):
                    continue
                name = (g.get('name') or '').strip()
                if not name:
                    continue
                parts = name.split(None, 1)
                first = parts[0]
                last = parts[1] if len(parts) > 1 else ''
                guest_type = (g.get('type') or '').strip().lower()
                if guest_type not in ('adult', 'child'):
                    guest_type = 'adult'
                RSVPGuest.objects.create(
                    rsvp=submission,
                    first_name=first,
                    last_name=last,
                    guest_type=guest_type,
                )
                stored_guests += 1

            for event_name in top_events:
                evt = str(event_name).strip()
                if evt:
                    RSVPEventSelection.objects.create(rsvp=submission, event_name=evt)

            invited.rsvped = True
            invited.current_guests = stored_guests
            invited.save(update_fields=['rsvped', 'current_guests'])
    except Exception:
        logger.exception("RSVP submit failed")
        return JsonResponse({'ok': False, 'error': 'Could not save RSVP'}, status=500)

    # Email is best-effort — never let it break the saved-RSVP response.
    try:
        _send_rsvp_emails(submission, payload)
    except Exception as e:
        logger.exception("RSVP email send failed: %s", e)

    return JsonResponse({'ok': True, 'id': submission.id})


def _send_rsvp_emails(submission, payload):
    invited = submission.invited_guest
    contact = payload.get('contact') or {}
    contact_email = (contact.get('email') or '').strip()
    contact_phone = (contact.get('phone') or '').strip()
    notes = (payload.get('notes') or '').strip()
    household_name = (payload.get('household') or '').strip() or f'{invited.first_name} {invited.last_name}'
    submitted_at = (payload.get('submittedAt') or '').strip() or submission.submitted_at.isoformat()
    guests = payload.get('guests') or []
    top_events = payload.get('topEventSelections') or []

    # Walk the party once: build the per-guest summary lines and capture the
    # first guest's first name (used to greet the confirmation recipient).
    guest_lines = []
    first_guest_name = ''
    for g in guests:
        if not isinstance(g, dict):
            continue
        name = (g.get('name') or '').strip()
        if not name:
            continue
        if not first_guest_name:
            first_guest_name = name.split()[0]
        gtype = (g.get('type') or '').strip()
        gevents = ', '.join(g.get('events') or [])
        line = f"  - {name}"
        if gtype:
            line += f" ({gtype})"
        if gevents:
            line += f" — events: {gevents}"
        guest_lines.append(line)
    if not first_guest_name:
        first_guest_name = invited.first_name

    summary_text = (
        f"Household: {household_name}\n\n"
        f"Guests:\n"
        + ('\n'.join(guest_lines) if guest_lines else '  (none)')
        + "\n\n"
        f"Top-level events: {', '.join(top_events) or '(none)'}\n\n"
        f"Contact:\n"
        f"  Email: {contact_email}\n"
        f"  Phone: {contact_phone or '—'}\n\n"
        f"Notes:\n{notes or '—'}\n"
    )

    notify_addr = getattr(settings, 'NOTIFICATION_EMAIL', '')

    # 1. Notification to the wedding couple
    if notify_addr:
        notification_text = (
            f"New RSVP received!\n\n"
            f"Submitted: {submitted_at}\n\n"
            f"{summary_text}\n"
            f"---\n"
            f"Sent from bermeawedding2026.com\n"
        )
        try:
            send_mail(
                subject='New RSVP — ' + household_name,
                message=notification_text,
                from_email=settings.DEFAULT_FROM_EMAIL,
                recipient_list=[notify_addr],
                fail_silently=False,
            )
        except Exception as e:
            logger.exception("Notification email failed: %s", e)

    # 2. Confirmation to guest (only if their address differs from the couple's)
    if contact_email and contact_email != notify_addr:
        confirmation_text = (
            f"Hi {first_guest_name},\n\n"
            f"Thank you for your RSVP! We're so excited to celebrate with you.\n\n"
            f"Here's a summary of your RSVP:\n\n"
            f"{summary_text}\n"
            f"Wedding Details:\n"
            f"  Date: Saturday, September 5, 2026\n"
            f"  Time: Ceremony at 5:30 PM | Doors open at 5:00 PM\n"
            f"  Venue: The Skyline\n"
            f"  Address: 707 Dawson St, San Antonio, TX 78202\n\n"
            f"Dress Code: Semi-Formal (required)\n"
            f"Hashtag: #BermeaEverAfter\n\n"
            f"If you need to make any changes, please contact us at:\n"
            f"bermeawedding@outlook.com\n\n"
            f"We can't wait to see you!\n\n"
            f"With love,\n"
            f"Brian & Aisha\n"
        )
        try:
            send_mail(
                subject='Your RSVP is confirmed — Brian & Aisha Wedding',
                message=confirmation_text,
                from_email=settings.DEFAULT_FROM_EMAIL,
                recipient_list=[contact_email],
                fail_silently=False,
            )
        except Exception as e:
            logger.exception("Confirmation email failed: %s", e)


# ----------------------------- Dashboard ------------------------------

def dashboard_login(request):
    if request.user.is_authenticated:
        return redirect('wedding:dashboard_home')
    error = ''
    if request.method == 'POST':
        username = (request.POST.get('username') or '').strip()
        password = request.POST.get('password') or ''
        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            return redirect('wedding:dashboard_home')
        error = 'Invalid username or password.'
    return render(request, 'dashboard/login.html', {'error': error})


def dashboard_logout(request):
    logout(request)
    return redirect('wedding:dashboard_login')


@login_required(login_url=DASHBOARD_LOGIN_URL)
def dashboard_home(request):
    invited_total  = InvitedGuest.objects.aggregate(s=Sum('max_guests'))['s'] or 0
    households     = InvitedGuest.objects.count()
    rsvped         = InvitedGuest.objects.filter(rsvped=True).count()
    attending      = InvitedGuest.objects.filter(rsvped=True).aggregate(s=Sum('current_guests'))['s'] or 0
    not_responded  = InvitedGuest.objects.filter(rsvped=False).count()
    return render(request, 'dashboard/home.html', {
        'invited_total': invited_total,
        'households': households,
        'rsvped': rsvped,
        'attending': attending,
        'not_responded': not_responded,
        'rsvp_deadline': 'July 24, 2026',
        'active_page': 'home',
    })


@login_required(login_url=DASHBOARD_LOGIN_URL)
def dashboard_guests(request):
    qs = InvitedGuest.objects.all().order_by('last_name', 'first_name')
    search = (request.GET.get('search') or '').strip()
    rsvped_filter = (request.GET.get('rsvped') or '').strip()
    if search:
        qs = qs.filter(
            Q(first_name__icontains=search) |
            Q(last_name__icontains=search) |
            Q(email__icontains=search)
        )
    if rsvped_filter == 'yes':
        qs = qs.filter(rsvped=True)
    elif rsvped_filter == 'no':
        qs = qs.filter(rsvped=False)
    guests = list(qs)
    return render(request, 'dashboard/guests.html', {
        'guests': guests,
        'search': search,
        'rsvped_filter': rsvped_filter,
        'total': len(guests),
        'active_page': 'guests',
    })


def _save_guest_from_form(post, guest=None):
    errors = {}
    first_name   = (post.get('first_name') or '').strip()
    last_name    = (post.get('last_name') or '').strip()
    email        = (post.get('email') or '').strip()
    phone        = (post.get('phone') or '').strip()
    partner_name = (post.get('partner_name') or '').strip()
    notes        = (post.get('notes') or '').strip()
    try:
        max_guests = int(post.get('max_guests') or 0)
    except (TypeError, ValueError):
        max_guests = 0
    if not first_name:
        errors['first_name'] = 'First name is required.'
    if not last_name:
        errors['last_name'] = 'Last name is required.'
    if max_guests < 1:
        errors['max_guests'] = 'Max guests must be at least 1.'
    if errors:
        return False, errors, guest
    if guest is None:
        guest = InvitedGuest()
    guest.first_name   = first_name
    guest.last_name    = last_name
    guest.email        = email or None
    guest.phone        = phone
    guest.partner_name = partner_name
    guest.max_guests   = max_guests
    guest.notes        = notes
    try:
        guest.save()
    except Exception as e:
        errors['__all__'] = str(e)
        return False, errors, guest
    return True, {}, guest


@login_required(login_url=DASHBOARD_LOGIN_URL)
def dashboard_guest_add(request):
    if request.method == 'POST':
        ok, errors, guest = _save_guest_from_form(request.POST)
        if ok:
            messages.success(request, 'Guest "%s %s" added.' % (guest.first_name, guest.last_name))
            return redirect('wedding:dashboard_guests')
        return render(request, 'dashboard/guest_form.html', {
            'guest': None, 'errors': errors, 'data': request.POST,
            'mode': 'add', 'active_page': 'add',
        })
    return render(request, 'dashboard/guest_form.html', {
        'guest': None, 'data': {'max_guests': 2},
        'mode': 'add', 'active_page': 'add',
    })


@login_required(login_url=DASHBOARD_LOGIN_URL)
def dashboard_guest_edit(request, pk):
    guest = get_object_or_404(InvitedGuest, pk=pk)
    if request.method == 'POST':
        ok, errors, _ = _save_guest_from_form(request.POST, guest=guest)
        if ok:
            messages.success(request, 'Guest "%s %s" updated.' % (guest.first_name, guest.last_name))
            return redirect('wedding:dashboard_guests')
        return render(request, 'dashboard/guest_form.html', {
            'guest': guest, 'errors': errors, 'data': request.POST,
            'mode': 'edit', 'active_page': 'guests',
        })
    return render(request, 'dashboard/guest_form.html', {
        'guest': guest,
        'data': {
            'first_name':   guest.first_name,
            'last_name':    guest.last_name,
            'email':        guest.email or '',
            'phone':        guest.phone or '',
            'partner_name': guest.partner_name or '',
            'max_guests':   guest.max_guests,
            'notes':        guest.notes or '',
        },
        'mode': 'edit', 'active_page': 'guests',
    })


@require_POST
@login_required(login_url=DASHBOARD_LOGIN_URL)
def dashboard_guest_delete(request, pk):
    guest = get_object_or_404(InvitedGuest, pk=pk)
    name = '%s %s' % (guest.first_name, guest.last_name)
    # RSVPSubmission(s) cascade-delete via FK on_delete=CASCADE
    guest.delete()
    messages.success(request, 'Guest "%s" deleted.' % name)
    return redirect('wedding:dashboard_guests')


@login_required(login_url=DASHBOARD_LOGIN_URL)
def dashboard_import(request):
    context = {'active_page': 'import', 'results': None}
    if request.method == 'POST':
        upload = request.FILES.get('file')
        if not upload:
            context['error'] = 'Please choose a file to upload.'
        else:
            file_name = (upload.name or '').lower()
            if not (file_name.endswith('.csv') or file_name.endswith('.xlsx')):
                context['error'] = 'Unsupported file type. Use .csv or .xlsx.'
            else:
                rows, parse_errors = _import_guests_file(upload, file_name)
                if parse_errors and not rows:
                    context['error'] = ' | '.join(parse_errors)
                else:
                    results = _apply_import_rows(rows)
                    results['errors'] = list(parse_errors) + results['errors']
                    context['results'] = results
    return render(request, 'dashboard/import.html', context)


@login_required(login_url=DASHBOARD_LOGIN_URL)
def dashboard_rsvps(request):
    qs = (
        RSVPSubmission.objects
        .select_related('invited_guest')
        .prefetch_related('guests', 'event_selections')
        .annotate(
            adult_count=Count('guests', filter=Q(guests__guest_type='adult')),
            child_count=Count('guests', filter=Q(guests__guest_type='child')),
        )
        .order_by('-submitted_at')
    )
    search = (request.GET.get('search') or '').strip()
    if search:
        qs = qs.filter(
            Q(invited_guest__first_name__icontains=search) |
            Q(invited_guest__last_name__icontains=search) |
            Q(email__icontains=search)
        )
    rsvps = list(qs)
    return render(request, 'dashboard/rsvps.html', {
        'rsvps': rsvps,
        'search': search,
        'total': len(rsvps),
        'active_page': 'rsvps',
    })


@login_required(login_url=DASHBOARD_LOGIN_URL)
def dashboard_rsvp_detail(request, pk):
    rsvp = get_object_or_404(
        RSVPSubmission.objects
            .select_related('invited_guest')
            .prefetch_related('guests', 'event_selections'),
        pk=pk,
    )
    return render(request, 'dashboard/rsvp_detail.html', {
        'rsvp': rsvp,
        'active_page': 'rsvps',
    })


@require_POST
def dashboard_rsvp_delete(request, pk):
    if not request.user.is_authenticated:
        return redirect('wedding:dashboard_login')

    submission = get_object_or_404(RSVPSubmission, pk=pk)

    invited_guest = submission.invited_guest

    # Delete the submission (cascades to RSVPGuest and RSVPEventSelection)
    submission.delete()

    # Reset the InvitedGuest RSVP status only if no other submissions remain
    remaining = RSVPSubmission.objects.filter(invited_guest=invited_guest).count()
    if remaining == 0:
        invited_guest.rsvped = False
        invited_guest.current_guests = 0
        invited_guest.save(update_fields=['rsvped', 'current_guests'])

    messages.success(
        request,
        'RSVP for %s %s has been deleted. They can now re-RSVP.' % (
            invited_guest.first_name, invited_guest.last_name
        ),
    )
    return redirect('wedding:dashboard_rsvps')


# ----------------------------- Import helpers -------------------------

def _import_guests_file(file_obj, file_name):
    """Parse an uploaded .csv/.xlsx into a list of normalized row dicts.

    Returns (rows, errors). On unrecoverable parse failure rows is [] and
    errors is non-empty.
    """
    rows = []
    errors = []

    try:
        if file_name.endswith('.xlsx'):
            if openpyxl is None:
                return [], ['openpyxl is not installed; cannot read .xlsx files.']
            wb = openpyxl.load_workbook(file_obj, read_only=True, data_only=True)

            # Walk every sheet until we find one with a recognizable header row.
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                all_rows = list(ws.iter_rows(values_only=True))
                if not all_rows:
                    continue

                header_row_idx = None
                headers = []
                for i, row in enumerate(all_rows):
                    row_lower = [str(c).lower().strip() if c else '' for c in row]
                    if any(x in row_lower for x in ['first_name', 'first name', 'guest 1 name', 'guest1 name']):
                        header_row_idx = i
                        headers = row_lower
                        break

                if header_row_idx is None:
                    continue

                for row in all_rows[header_row_idx + 1:]:
                    if not any(row):
                        continue
                    cells = [str(c).strip() if c is not None else '' for c in row]
                    parsed = _parse_row(headers, cells)
                    if parsed:
                        rows.append(parsed)

                if rows:
                    break

        elif file_name.endswith('.csv'):
            content = file_obj.read()
            try:
                text = content.decode('utf-8-sig')
            except UnicodeDecodeError:
                text = content.decode('latin-1')

            reader = csv.reader(io.StringIO(text))
            all_rows = list(reader)
            if not all_rows:
                return [], ['Empty CSV file']

            headers = [h.lower().strip() for h in all_rows[0]]
            for row in all_rows[1:]:
                if not any((c or '').strip() for c in row):
                    continue
                cells = [(c or '').strip() for c in row]
                parsed = _parse_row(headers, cells)
                if parsed:
                    rows.append(parsed)
        else:
            return [], ['Unsupported file type. Use .csv or .xlsx.']

    except Exception as e:
        return [], ['Error reading file: %s' % e]

    return rows, errors


def _parse_row(headers, cells):
    """Map a single row of cells onto our canonical guest dict.

    Header lookup is substring-based and case-insensitive so the same parser
    handles both compact ("first_name") and human ("Guest 1 Name") sheets.
    Returns None for rows that have no usable first name.
    """
    def get(col_options):
        for opt in col_options:
            for i, h in enumerate(headers):
                if opt in h and i < len(cells):
                    val = cells[i].strip()
                    if val and val.lower() not in ('nan', 'none', 'n/a'):
                        return val
        return ''

    first_name = get(['first_name', 'first name'])
    last_name  = get(['last_name', 'last name'])

    if not first_name and not last_name:
        full = get(['guest 1 name', 'guest1 name', 'full_name', 'full name', 'name'])
        if full:
            parts = full.strip().split(' ', 1)
            first_name = parts[0]
            last_name = parts[1] if len(parts) > 1 else ''

    if not first_name:
        return None

    max_g_raw = get(['max_guests', 'max guests', 'total number of guests', 'guests'])
    try:
        max_guests = max(1, int(float(max_g_raw))) if max_g_raw else 2
    except (TypeError, ValueError):
        max_guests = 2

    # Accept either a combined "Partner Name" column or partner_first/_last.
    partner_full = get(['partner_full_name', 'partner full name', 'partner name'])
    if not partner_full:
        pf = get(['partner_first', 'partner first'])
        pl = get(['partner_last', 'partner last'])
        partner_full = (pf + ' ' + pl).strip()

    return {
        'first_name':   first_name,
        'last_name':    last_name,
        'email':        get(['email', 'email address']),
        'phone':        get(['phone', 'phone number']),
        'max_guests':   max_guests,
        'partner_name': partner_full,
        'notes':        get(['notes', 'note']),
    }


def _apply_import_rows(rows):
    """Persist parsed rows. Match on email, then first+last name, then first
    name alone if last name is empty. RSVP state is never touched.
    """
    results = {'created': 0, 'updated': 0, 'skipped': 0, 'errors': []}

    for row in rows:
        try:
            first = (row.get('first_name') or '').strip()
            last  = (row.get('last_name') or '').strip()
            email = (row.get('email') or '').strip()
            phone = (row.get('phone') or '').strip()
            partner_name = (row.get('partner_name') or '').strip()
            max_g = row.get('max_guests') or 1

            if not first:
                results['skipped'] += 1
                continue

            existing = None
            if email:
                existing = InvitedGuest.objects.filter(email__iexact=email).first()
            if not existing and last:
                existing = InvitedGuest.objects.filter(
                    first_name__iexact=first,
                    last_name__iexact=last,
                ).first()
            if not existing and not last:
                existing = InvitedGuest.objects.filter(first_name__iexact=first).first()

            if existing:
                changed = False
                if max_g and existing.max_guests != max_g:
                    existing.max_guests = max_g
                    changed = True
                if email and not existing.email:
                    existing.email = email
                    changed = True
                if phone and not existing.phone:
                    existing.phone = phone
                    changed = True
                if partner_name and not existing.partner_name:
                    existing.partner_name = partner_name
                    changed = True
                if changed:
                    existing.save()
                    results['updated'] += 1
                else:
                    results['skipped'] += 1
            else:
                InvitedGuest.objects.create(
                    first_name=first,
                    last_name=last,
                    email=(email or None),
                    phone=phone,
                    partner_name=partner_name,
                    max_guests=max_g,
                    notes=(row.get('notes') or ''),
                )
                results['created'] += 1

        except Exception as e:
            results['errors'].append('%s %s: %s' % (
                row.get('first_name', ''), row.get('last_name', ''), e))

    return results
